import allure
import openpyxl
import pytest
from allure_commons.types import AttachmentType
from appium import webdriver
from appium.webdriver.appium_service import AppiumService
import time
import sys

from appium import webdriver
from pathlib import Path
# from appium.webdriver.common.appiumby import AppiumBy
# from appium.webdriver.appium_service import AppiumService
# from selenium.webdriver.common.keys import Keys
from appium.webdriver.common.appiumby import AppiumBy
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
# from scrollUtility import ScrollUtility

from appium import webdriver
# from appium.webdriver.common.appiumby import AppiumBy
# from appium.webdriver.appium_service import AppiumService
# from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import random
from selenium.webdriver.support.select import Select

from appium import webdriver
from pathlib import Path
from appium.webdriver.common.appiumby import AppiumBy
from appium.webdriver.appium_service import AppiumService
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from appium.webdriver.common.touch_action import TouchAction
from utilities import dataProvider
"""FOr Hard COde data get from CSV"""
def getData():
    workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")
    sheet = workbook["Sheet1"]
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    mainList = []

    for rows in range(2, total_rows + 1):
        dataList = []
        for cols in range(1, total_cols + 1):
            data = sheet.cell(row=rows, column=cols).value
            dataList.insert(cols, data)
        mainList.insert(rows, dataList)
    return mainList
####################################

@pytest.mark.usefixtures("log_on_failure")
@pytest.mark.parametrize("city,country", getData())
def test_login(city, country, appium_driver):
    driver = appium_driver
    driver.find_element(By.XPATH, "//android.widget.TextView[@text='Hotels']").click()
    time.sleep(2)
    driver.find_element(By.XPATH, "//android.widget.TextView[@text='City, Area or Property Name']").click()
    time.sleep(2)
    print(city)
    driver.find_element(By.ID, "com.goibibo:id/edtSearch").send_keys(city)
    time.sleep(2)
    driver.find_elements(By.ID, "com.goibibo:id/lytChildNode1")[0].click()
    time.sleep(4)
    driver.find_element(By.ID, "com.goibibo:id/verticalHomeSearchButton").click()
    time.sleep(10)
    citytext = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Top Areas to Stay')]").text
    print(citytext)
    """Attach Screenshot to allure report"""
    allure.attach(driver.get_screenshot_as_png(), name="Screenshot",
                  attachment_type=AttachmentType.PNG)  ###Success Sceenshot
