import openpyxl

path = "..//excel//AbisServerInfor.xlsx"
sheetname = "Sheet1"


def getRowCount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_row


def getColumnCount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_column


def getCellData(path, sheetname, rowNum, colNum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.cell(row=rowNum, column=colNum).value


def setCellData(path, sheetname, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(path)


# path ="..//excel//AbisServerInfor.xlsx"
# sheetname= "Sheet1"

rows = getRowCount(path, sheetname)
cols = getColumnCount(path, sheetname)

print(rows, "   ", cols)
print(getCellData(path,sheetname,2,3))
setCellData(path,sheetname,1,5,"Hello")
print(getCellData(path,sheetname,1,5))