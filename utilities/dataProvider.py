import openpyxl


def getData(sheetName):
    workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")
    sheet = workbook[sheetName]
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    mainList = []

    for rows in range(2, total_rows + 1):
        dataList = []
        for cols in range(1, total_cols + 1):
            data = sheet.cell(row=rows, column=cols).value
            dataList.insert(cols, data)
        mainList.insert(rows, dataList)
    return mainList