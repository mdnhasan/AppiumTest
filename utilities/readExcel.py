import openpyxl

# workbook = openpyxl.load_workbook("D:\\WORK\\AbisServerInfor.xlsx")
workbook = openpyxl.load_workbook("..//excel//AbisServerInfor.xlsx")
sheet = workbook["Sheet1"]

totalrows = sheet.max_row
totalcolumn = sheet.max_column

# print("total rows are", str(totalrows),"Total cols are", str(totalcolumn))
#
# print(sheet.cell(row=2, column=2).value)

for rows in range(1, totalrows+1):
    for cols in range(1, totalcolumn+1):
        print(sheet.cell(row=rows,column=cols).value, end="   ")
    print()