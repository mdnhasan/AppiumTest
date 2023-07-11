import openpyxl

# workbook = openpyxl.load_workbook("D:\\WORK\\AbisServerInfor.xlsx")
workbook = openpyxl.load_workbook("..//excel//AbisServerInfor.xlsx")

sheet = workbook["Sheet1"]

# sheet.cell(row=1, column=4).value="BranchName"
# workbook.save("..//excel//AbisServerInfor.xlsx")

for rows in range(2,6):
    for cols in range(4,5):
        sheet.cell(row=rows, column=cols).value = "NewBranch"
workbook.save(r"..//excel//AbisServerInfor.xlsx")
